from flask import Flask, render_template, request, redirect, url_for, Response, send_file
import sqlite3
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
DB_NAME = "pct.db"

DISEASES = [
    "Sepsis","Stroke","STEMI","Head Injury","Pneumonia","COPD",
    "Asthma","DM","HT","CKD","DHF","Diarrhea",
    "ภาวะซีดในหญิงตั้งครรภ์","PIH","PPH",
    "Birth Asphyxia","AIDS","TB","Teenage Pregnancy","ฟันผุ"
]

MONTHS = ["ม.ค.","ก.พ.","มี.ค.","เม.ย.","พ.ค.","มิ.ย.","ก.ค.","ส.ค.","ก.ย.","ต.ค.","พ.ย.","ธ.ค."]

SEPSIS_ITEMS = [
    "Lactate",
    "H/C ก่อนให้ ATB",
    "ATB ภายใน 1 ชม.",
    "Fluid Resuscitate",
    "ใส่สายสวนปัสสาวะ",
    "NE (septic shock)",
    "การส่งต่อทันเวลา",
    "อัตราตาย Sepsis",
    "Severe Sepsis ระหว่างดูแล",
    "Sepsis จากบ้าน"
]
STROKE_ITEMS = [
    "จำนวนผู้ป่วย stroke ทั้งหมด",
    "จำนวนผู้ป่วย stroke เพศชาย",
    "จำนวนผู้ป่วย stroke เพศหญิง",
    "จำนวนผู้ป่วย stroke เสียชีวิต",
    "ผู้รับบริการอายุ <= 40 ปี",
    "ผู้รับบริการอายุ 40 - 50 ปี",
    "ผู้รับบริการอายุ 51 -60 ปี",
    "ผู้รับบริการอายุ 61 - 70 ปี",
    "ผู้รับบริการอายุ >= 71 ปี",
    "มีประวัติเป็น DM",
    "มีประวัติเป็น HT",
    "มีประวัติเป็น DMและHT",
    "มีประวัติเป็น AF",
    "มีประวัติเป็นโรคอื่น ๆ",
    "ไม่มีโรคร่วม",
    "มารับบริการช่วงเวลา 00.30-08.30 น.",
    "มารับบริการช่วงเวลา 08.30-16.30 น.",
    "มารับบริการช่วงเวลา 16.30-00.30 น.",
    "จำนวนผู้ป่วย Stroke fast track",
    "ร้อยละของผู้ป่วย (onset to ER ภายใน 2 ชม.)",
    "ผู้ป่วย NCD CLINIC",
    "ผู้ป่วยทั่วไป",
    "ร้อยละของผู้ป่วย (onset to ER ภายใน 3 ชม.)",
    "ผู้ป่วย NCD CLINIC",
    "ผู้ป่วยทั่วไป",
    "ร้อยละของผู้ป่วย (onset to ER ภายใน 4.5 ชม.)",
    "ผู้ป่วย NCD CLINIC",
    "ผู้ป่วยทั่วไป",
    "ร้อยละของผู้ป่วย (onset to ER ≥ 4.5 ชม.)",
    "ร้อยละผู้ป่วย Stroke ที่มาด้วยระบบบริการ EMS",
    "ร้อยละผู้ป่วย Stroke ที่ได้รับการดูแลตามแนวทางที่กำหนด",
    "ร้อยละผู้ป่วยที่ Door to ER ภายใน 30 นาที",
    "ร้อยละผู้ป่วย Stroke fast track ที่ Door to ER ภายใน 30 นาที",
    "ร้อยละผู้ป่วยที่ Door to Refer ภายใน 80 นาที",
    "ร้อยละผู้ป่วย Stroke fast track ที่ Door to Refer ภายใน 80 นาที"
]
HEAD_INJURY_ITEMS = [
    "จำนวนผู้ป่วย HI ทั้งหมด",
    "จำนวนผู้ป่วย Mild Head injury with Low Risk",
    "จำนวนผู้ป่วย Mild Head injury with Moderate Risk",
    "จำนวนผู้ป่วย Mild Head Injury with High Risk",
    "จำนวนผู้ป่วย Moderate Head Injury",
    "จำนวนผู้ป่วย Severe Head injury",
    "จำนวนผู้ป่วย Severe Head injury เสียชีวิต ณ จุดเกิดเหตุ",
    "ร้อยละผู้ป่วย Head Injury ที่มาด้วยระบบบริการการแพทย์ฉุกเฉิน",
    "ร้อยละความสามารถในการตอบสนองต่อการเรียกใช้ EMS ภายใน 10 นาที กรณีผู้ป่วยอุบัติเหตุจราจร",
    "ร้อยละการประเมินจัดระดับความรุนแรงมีความถูกต้อง",
    "ผู้ป่วย Mild Head injury with Moderate Risk ได้ admit ,refer",
    "ผู้ป่วย Mild Head Injury with High Risk ได้ refer",
    "ผู้ป่วย Moderate Head Injury ได้ refer",
    "ผู้ป่วย Severe Head injury ได้รับการ refer",
    "ผู้ป่วย Severe Head injury เสียชีวิต ที่ ER",
    "อัตราการเสียชีวิตจากอุบัติเหตุจราจร",
    "Door To Refer ภายในเวลา 1 ชั่วโมง"
]
STEMI_ITEMS = [
    "จำนวนผู้ป่วย STEMI ทั้งหมด",
    "EKG ภายใน 5 นาที",
    "Door to doctor ภายใน 10 นาที",
    "เจาะ Trop-I < 15 นาที",
    "consult < 10 นาที",
    "ภายใน 10 นาทีหลังการวินิจฉัย",
    "door to ER < 30 นาที (กรณีไม่ให้ SK)",
    "door to refer < 80 นาที (กรณีไม่ให้ SK)",
    "ทำ PCI สำเร็จ",
    "จำนวนผู้ป่วยที่ได้รับ SK ทั้งหมด",
    "Door to SK Time ภายใน 30 นาที",
    "มีอาการเจ็บหน้าอกน้อยกว่า 3 ชม.ก่อนมาโรงพยาบาล",
    "ประวัติ HT",
    "ประวัติ DM",
    "ประวัติ DM-HT",
    "ผู้ป่วยมีโรคร่วมอื่นๆ",
    "ไม่มีโรคร่วม",
    "ประวัติสูบบุหรี่",
    "ได้รับการรักษาตามมาตรฐาน",
    "ใช้ MOVIE MP2 AAA",
    "ได้รับการทบทวน case",
    "อุบัติการณ์เสียชีวิตในหน่วยงาน",
    "อุบัติการณ์เสียชีวิตขณะส่งต่อ",
    "อุบัติการณ์เสียชีวิตใน IPD",
    "อุบัติการณ์เสียชีวิตก่อนถึง ร.พ.",
    "เรียกใช้บริการ 1669",
    "มี APIE",
    "หลัง D/C ได้รับการเยี่ยมบ้าน",
    "อัตราการเกิดภาวะแทรกซ้อนโรคหัวใจและหลอดเลือดในผู้ป่วย HT",
    "มีอาการเจ็บหน้าอกน้อยกว่า 60 นาที"
]
PNEUMONIA_ITEMS = [
    "จำนวนผู้ป่วย Pneumonia ที่ Admitted ทั้งหมด",
    "อุบัติการณ์การเกิด Respiratory failure",
    "อัตราการกลับมารักษาซ้ำ Re–admitted ภายใน 28 วัน",
    "ร้อยละของผู้ดูแลผู้ป่วยที่สามารถเคาะปอดดูดเสมหะและฝึกหายใจได้เอง",
    "ร้อยละของผู้ป่วยและครอบครัวได้รับการวางแผนจำหน่าย",
    "จำนวนผู้ป่วย Pneumonia ที่ refer ทั้งหมด"
]
DIARRHEA_ITEMS = [
    "จำนวนผู้ป่วย Diarrhea ทั้งหมด (ครั้ง/คน)",
    "อุบัติการณ์ของผู้ป่วยที่มีภาวะ Shock จากการขาดน้ำ",
    "ร้อยละผู้ป่วยที่มาด้วยภาวะเกลือแร่ผิดปกติได้รับการแก้ไข",
    "ร้อยละผู้ป่วยได้รับคำแนะนำเรื่องการรับประทานอาหาร/โรค",
    "อุบัติการณ์ของผู้ป่วยเสียชีวิตด้วยโรคอุจจาระร่วง",
    "อุบัติการณ์ของผู้ป่วยที่ส่งต่อด้วย Diarrhea C Septic Shock"
]
DHF_ITEMS = [
    "ผู้ป่วยที่ได้รับการวินิจฉัย DF",
    "ผู้ป่วยที่ได้รับการวินิจฉัย DHF",
    "รวมผู้ป่วยทั้งหมดที่ได้รับการวินิจฉัย DF/ DHF",
    "อุบัติการณ์ผู้ป่วยที่มีภาวะ Shock ก่อนการรักษาใน รพ",
    "อุบัติการณ์ผู้ป่วยที่มีภาวะ Shock ขณะรับการรักษาใน รพ",
    "จำนวนผู้ป่วยที่มีภาวะแทรกซ้อนและรับการส่งต่อ",
    "อัตราการส่งต่อตามแนวทางปฏิบัติ",
    "อัตราการตาย"
]
COPD_ITEMS = [
    "จำนวนผู้ป่วย COPD ทั้งหมด",
    "ร้อยละผู้ป่วย COPD ใช้ยาพ่นได้ถูกต้อง",
    "อัตราการ Re-admitted ภายใน 28 วัน",
    "อัตราการ Re-visited ภายใน 48 ชั่วโมง",
    "อัตราการเกิด Exacerbation ที่ ER",
    "อัตราผู้ป่วยมารับบริการก่อนนัด",
    "อัตราผู้ป่วยขาดนัด"
]
ASTHMA_ITEMS = [
    "จำนวนผู้ป่วย Asthma ทั้งหมด",
    "ผู้ป่วยได้รับความรู้เรื่องโรคและปฏิบัติตัวได้ถูกต้อง",
    "ร้อยละผู้ป่วย Asthma ได้รับยา ICS",
    "อัตราผู้ป่วย Asthma ใช้ยาพ่นได้ถูกต้อง",
    "อัตราการ Re-admitted ภายใน 28 วัน",
    "อัตราการ Re-visited ภายใน 48 ชั่วโมง",
    "อัตราผู้ป่วยมารับบริการก่อนนัด"
]
DM_ITEMS = [
    "ร้อยละของประชากรอายุ 35 ปีขึ้นไปที่ได้รับการคัดกรองเพื่อวินิจฉัยเบาหวาน",
    "ร้อยละของประชากรอายุ 35-59 ปีที่ได้รับการคัดกรองเพื่อวินิจฉัยเบาหวาน",
    "อัตราประชากรกลุ่มเสี่ยงเบาหวานในพื้นที่รับผิดชอบของปีที่ผ่านมาได้รับการตรวจน้ำตาลซ้ำ",
    "อัตราผู้ป่วยเบาหวานรายใหม่จากกลุ่มเสี่ยงปีที่ผ่านมา",
    "ร้อยละการตรวจติดตามยืนยันวินิจฉัยกลุ่มสงสัยป่วยเบาหวาน",
    "อัตราผู้ป่วยเบาหวานรายใหม่จากกลุ่มสงสัยป่วยที่ได้รับการตรวจยืนยัน",
    "ร้อยละของผู้ป่วยเบาหวานรายใหม่ลดลง",
    "ร้อยละผู้ป่วยเบาหวานที่ขึ้นทะเบียน และได้รับการตรวจติดตามและรักษาที่เหมาะสม",
    "ร้อยละผู้ป่วยเบาหวานที่มีภาวะอ้วนลงพุง",
    "ร้อยละของผู้ป่วยเบาหวานที่มีภาวะอ้วนลงพุง ลดลงจากงบประมาณที่ผ่านมา",
    "ร้อยละของผู้ป่วยโรคเบาหวานที่ได้รับการตรวจ HbA1c อย่างน้อย 1 ครั้ง/ปี",
    "ร้อยละผู้ป่วยโรคเบาหวานที่ควบคุมระดับน้ำตาลได้ดี",
    "ร้อยละของผู้ป่วยเบาหวานที่มีความดันโลหิตควบคุมได้ตามเกณฑ์",
    "อัตราผู้ป่วยเบาหวานที่ได้รับการตรวจไขมัน LDL",
    "อัตราผู้ป่วยเบาหวานที่ได้รับการตรวจไขมัน LDL และมีค่า LDL < 100 mg/dl",
    "ร้อยละของผู้ป่วยเบาหวานที่ได้รับการตรวจภาวะแทรกซ้อนทางตา",
    "ร้อยละของผู้ป่วยเบาหวานที่ได้รับการตรวจภาวะแทรกซ้อนทางเท้า",
    "อัตราการเกิดภาวะแทรกซ้อนเฉียบพลันในผู้ป่วยเบาหวาน"
]
HT_ITEMS = [
    "ร้อยละของประชากรอายุ 35 ปีขึ้นไปที่ได้รับการคัดกรองเพื่อวินิจฉัยความดันโลหิตสูง",
    "ร้อยละของประชากรอายุ 35-59 ปีขึ้นไปที่ได้รับการคัดกรองเพื่อวินิจฉัยความดันโลหิตสูง",
    "ร้อยละการตรวจติดตามยืนยันวินิจฉัยกลุ่มสงสัยป่วยโรคความดันโลหิตสูง",
    "ร้อยละของผู้ป่วยความดันโลหิตสูงรายใหม่จากกลุ่มสงสัยป่วยที่ได้รับการตรวจติดตาม",
    "ร้อยละของผู้ป่วยความดันโลหิตสูงรายใหม่จากผู้ที่มีระดับความดันโลหิตอยู่ในเกณฑ์เกือบสูง (กลุ่มเสี่ยงความดันโลหิตสูง)",
    "ร้อยละผู้ป่วยโรคความดันโลหิตสูงที่ควบคุมความดันโลหิตได้ดี",
    "ร้อยละของผู้ป่วยความดันโลหิตสูงรายใหม่ลดลง",
    "อัตราผู้ป่วยความดันโลหิตสูงที่ขึ้นทะเบียน และมารับการรักษาในเขตพื้นที่รับผิดชอบ",
    "อัตราการเกิดภาวะแทรกซ้อนโรคหลอดเลือดสมองในผู้ป่วย HT",
    "อัตราการเกิดภาวะแทรกซ้อนโรคหัวใจและหลอดเลือดในผู้ป่วย HT",
    "ร้อยละผู้ป่วย CVD risk ≥ 20% มีภาวะแทรกซ้อนโรคหลอดเลือดสมอง"
]
CKD_ITEMS = [
    "ร้อยละของผู้ป่วย DM และ/หรือ HT ที่ได้รับการค้นหาและคัดกรองโรคไตเรื้อรัง",
    "ร้อยละของผู้ป่วย DM และ/หรือ HT ที่เป็นผู้ป่วยโรคไตเรื้อรังรายใหม่",
    "ร้อยละของผู้ป่วย CKD ที่มี BP < 140/90 mmHg",
    "ร้อยละของผู้ป่วย CKD ที่ได้รับ ACEi/ARB",
    "ร้อยละของผู้ป่วย CKD ที่มีอัตราการลดลงของ eGFR < 5 ml/min/1.73m2/yr",
    "ร้อยละของผู้ป่วย CKD มีระดับ Hb > 10 gm/dl หรือ Hct > 30%",
    "ร้อยละผู้ป่วย CKD ที่มี HbA1c ระหว่าง 6.5-7.5 mg% (เฉพาะผู้ป่วย DM)",
    "ร้อยละของผู้ป่วย CKD กลุ่มเสี่ยงต่อโรคหลอดเลือดและหัวใจได้รับยากลุ่ม Statin (CKD stage 3-4 และมีอายุตั้งแต่ 50 ปี)",
    "ร้อยละของผู้ป่วย CKD มีค่า serum K < 5.5 mEq/L",
    "ร้อยละของผู้ป่วย CKD มีค่า serum HCO3 > 22 mEq/L",
    "ร้อยละผู้ป่วย CKD ได้รับการตรวจ urine protein"
]
ANEMIA_PREG_ITEMS = [
    "หญิงไทยรับบริการฝากครรภ์ได้รับการตรวจ Hct ทั้งหมด",
    "หญิงไทยรับบริการฝากครรภ์ได้รับการตรวจ Hct ผลอยู่ระหว่าง 1 - 32%",
    "ร้อยละของหญิงตั้งครรภ์ที่ได้รับยาเม็ดเสริมธาตุเหล็กเสริมไอโอดีน"
]
PIH_ITEMS = [
    "จำนวนมารดาที่มีภาวะ Severe features (>160/110 mmHg) ทั้งหมด",
    "จำนวนมารดาที่มีภาวะ Severe features (>160/110 mmHg) ไม่มีภาวะชักทั้งหมด",
    "จำนวนมารดาที่มีภาวะ Eclampsia (มีอาการชัก) ทั้งหมด",
    "มารดาที่มาคลอด BP ≥ 140/90 mmHg Clinical of SPE ได้รับการตรวจ Lab Toxemia",
    "มารดา Severe features ได้รับยากันชักก่อนนำส่งทุกราย",
    "อุบัติการณ์มารดามีอาการชักขณะดูแล/ชักขณะส่งต่อ",
    "มารดาที่มีอาการชัก ต้องได้รับการทบทวน Case ทุกราย",
    "อุบัติการณ์ มารดา Severe features คลอดเนื่องจาก Refer ไม่ทัน"
]
PPH_ITEMS = [
    "จำนวนมารดาคลอดทั้งหมด",
    "จำนวนมารดาตกเลือดหลังคลอดทั้งหมด",
    "อัตราการตกเลือด",
    "มีภาวะ Shock จากภาวะตกเลือดหลังคลอด",
    "อุบัติการณ์มารดาตกเลือดหลังคลอดเสียชีวิตขณะส่งต่อ",
    "อุบัติการณ์มารดาตกเลือดหลังคลอดเสียชีวิตในโรงพยาบาล",
    "มารดาตกเลือดหลังคลอดทุกรายได้รับการทบทวน",
    "จำนวนมารดาตกเลือดหลังคลอดภายใน 24 ชม.",
    "จำนวนมารดาตกเลือดหลังคลอดภายใน 24 ชม - 6 สัปดาห์"
]
BIRTH_ASPHYXIA_ITEMS = [
    "อัตราทารกแรกเกิดมีภาวะขาดออกซิเจน",
    "จำนวนทารกมีภาวะขาดออกซิเจนรุนแรงได้รับการส่งต่อ",
    "อุบัติการณ์จำนวนทารกที่เสียชีวิต"
]
AIDS_ITEMS = [
    "จำนวน PHA ที่ขึ้นทะเบียนทั้งหมด",
    "ร้อยละ PHA ที่ได้รับการรักษาด้วยยาต้านไวรัสทั้งหมด",
    "ร้อยละ PHA ที่ได้รับการรักษาด้วยยาต้านไวรัสติดต่อกันมากกว่า 6 เดือน มี Viral Load < 50 Copies"
]
TB_ITEMS = [
    "จำนวนผู้ป่วยวัณโรคทั้งหมด",
    "จำนวนผู้ป่วยวัณโรครายใหม่เสมหะบวก (NEW M+)",
    "อัตราการให้คำปรึกษาคัดกรอง HIV ใน TB ทั้งหมด (VCT)",
    "อัตราการยินยอมเจาะ HIV ใน TB ทั้งหมดหลังได้รับคำปรึกษา",
    "อัตราพบ HIV positive รายใหม่ในผู้ป่วย TB ทั้งหมด",
    "อัตราการเจาะ CD4 ในผู้ป่วย TB/HIV",
    "อัตราการให้ ARV ในผู้ป่วย TB/HIV ที่ CD4 < 250",
    "อัตราการแปรเปลี่ยนของเสมหะจากบวกเป็นลบในผู้ป่วยวัณโรครายใหม่เสมหะพบเชื้อ (Sputum conversion)",
    "อัตราความสำเร็จในการรักษาวัณโรคปอดรายใหม่เสมหะพบเชื้อ (Success rate)",
    "อัตราการขาดยาในผู้ป่วยวัณโรครายใหม่เสมหะพบเชื้อ (Default rate)",
    "อัตราการตายในผู้ป่วยวัณโรครายใหม่เสมหะพบเชื้อ (Death rate)",
    "อัตราการตรวจพบวัณโรคปอดรายใหม่เสมหะพบเชื้อ (Case detection rate)"
]
TEENAGE_ITEMS = [
    "อัตราการคลอดมีชีพของหญิง 15-19 ปี / พันประชากร",
    "อัตราการตั้งครรภ์ซ้ำในหญิงวัยรุ่น",
    "อัตราการคุมกำเนิดแบบกึ่งถาวรในหญิงตั้งครรภ์อายุ 15-19 ปี"
]
DENTAL_ITEMS = [
    "ร้อยละของเด็กอายุ 12 ปี ฟันดีไม่มีผุ (Cavity Free)"
]

def get_items_by_disease(disease):
    if disease == "Sepsis":
        return SEPSIS_ITEMS

    elif disease == "Stroke":
        return STROKE_ITEMS

    elif disease == "Head Injury":
        return HEAD_INJURY_ITEMS

    elif disease == "STEMI":
        return STEMI_ITEMS

    elif disease == "Pneumonia":
        return PNEUMONIA_ITEMS

    elif disease == "COPD":
        return COPD_ITEMS

    elif disease == "Asthma":
        return ASTHMA_ITEMS

    elif disease == "DM":
        return DM_ITEMS

    elif disease == "HT":
        return HT_ITEMS

    elif disease == "CKD":
        return CKD_ITEMS

    elif disease == "DHF":
        return DHF_ITEMS

    elif disease == "Diarhea":
        return DIARRHEA_ITEMS

    elif disease == "ภาวะซีดในหญิงตั้งครรภ์":
        return ANEMIA_PREG_ITEMS

    elif disease == "PIH":
        return PIH_ITEMS

    elif disease == "PPH":
        return PPH_ITEMS

    elif disease == "Birth Asphisia":
        return BIRTH_ASPHYXIA_ITEMS

    elif disease == "AID":
        return AIDS_ITEMS

    elif disease == "TB":
        return TB_ITEMS

    elif disease == "Teenage Pregnancy":
        return TEENAGE_ITEMS

    elif disease == "ฟันผุ":
         return DENTAL_ITEMS

    return []

def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS pct_values (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fiscal_year INTEGER NOT NULL,
            disease TEXT NOT NULL,
            indicator TEXT NOT NULL,
            month_index INTEGER NOT NULL,
            value INTEGER,
            created_at TEXT,
            updated_at TEXT,
            UNIQUE(fiscal_year, disease, indicator, month_index)
        )
    """)
    conn.commit()
    conn.close()
def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS pct_values (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fiscal_year INTEGER NOT NULL,
            disease TEXT NOT NULL,
            indicator TEXT NOT NULL,
            month_index INTEGER NOT NULL,
            value INTEGER,
            created_at TEXT,
            updated_at TEXT,
            UNIQUE(fiscal_year, disease, indicator, month_index)
        )
    """)
    conn.commit()
    conn.close()

init_db()
def get_values(disease, fiscal_year):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("""
        SELECT indicator, month_index, value
        FROM pct_values
        WHERE disease=? AND fiscal_year=?
    """, (disease, fiscal_year))
    rows = cur.fetchall()
    conn.close()

    values = {}
    for indicator, month_index, value in rows:
        values[(indicator, month_index)] = value
    return values



@app.route("/disease/<path:name>", methods=["GET", "POST"])
def disease(name):
    name = name.replace("_", " ")
    items = get_items_by_disease(name)
    fiscal_year = int(request.args.get("year", "2568"))

    if request.method == "POST":
        fiscal_year = int(request.form.get("fiscal_year", fiscal_year))
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = sqlite3.connect(DB_NAME)
        cur = conn.cursor()

        for item in items:
            for m in range(12):
                field_name = f"{item}_{m}"
                value = request.form.get(field_name)

                if value == "" or value is None:
                    value = None
                else:
                    value = int(value)

                cur.execute("""
                    INSERT INTO pct_values
                    (fiscal_year, disease, indicator, month_index, value, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(fiscal_year, disease, indicator, month_index)
                    DO UPDATE SET
                        value=excluded.value,
                        updated_at=excluded.updated_at
                """, (fiscal_year, name, item, m, value, now, now))

        conn.commit()
        conn.close()

        return redirect(url_for("disease", name=name.replace(" ", "_"), year=fiscal_year))

    values = get_values(name, fiscal_year)

    return render_template(
        "disease.html",
        name=name,
        months=MONTHS,
        items=items,
        values=values,
        fiscal_year=fiscal_year
    )

@app.route("/export_excel/<path:name>")
def export_excel(name):
    name = name.replace("_", " ")
    fiscal_year = int(request.args.get("year", "2568"))
    values = get_values(name, fiscal_year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name

    headers = ["ลำดับ", "ตัวชี้วัด"] + MONTHS
    ws.append(headers)

    for idx, item in enumerate(SEPSIS_ITEMS, start=1):
        row = [idx, item]
        for m in range(12):
            value = values.get((item, m), "")
            row.append("" if value is None else value)
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="0F8A5F")
    white_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CCCCCC")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35

    for col in range(3, 15):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{name}_{fiscal_year}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/export/<path:name>")
def export_report(name):
    name = name.replace("_", " ")
    fiscal_year = int(request.args.get("year", "2568"))
    values = get_values(name, fiscal_year)

    lines = []
    header = ["ลำดับ", "ตัวชี้วัด"] + MONTHS
    lines.append(",".join(header))

    for idx, item in enumerate(SEPSIS_ITEMS, start=1):
        row = [str(idx), item]
        for m in range(12):
            value = values.get((item, m), "")
            row.append("" if value is None else str(value))
        lines.append(",".join(row))

    csv_data = "\ufeff" + "\n".join(lines)
    filename = f"{name}_{fiscal_year}.csv"

    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.route("/")
def home():
    year = request.args.get("year", "2568")
    return render_template("home.html", diseases=DISEASES, year=year)

init_db()
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)